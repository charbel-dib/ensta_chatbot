from __future__ import annotations

import json
import re
import hashlib
from pathlib import Path
from typing import Any, Dict, List, Set, Optional, Tuple

from openpyxl import load_workbook


# ===== PATHS =====
PROJECT_ROOT = Path(__file__).resolve().parents[1]  # ENSTA_DATASET/
CONFIG_DIR = PROJECT_ROOT / "00_config"

# Racines à scanner (tu peux en mettre 2 comme demandé)
XLSX_ROOTS = [
    PROJECT_ROOT / "01_raw" / "Fiches cours FISE",
    PROJECT_ROOT / "01_raw" / "Fiches cours FIPA",
]

OUT_JSONL = PROJECT_ROOT / "02_clean" / "xlsx_courses_clean.jsonl"

# Paramètres extraction
ROW_DESC_1 = 12
ROW_DESC_2 = 13

# Logs
PRINT_EVERY = 50

SEM_RE = re.compile(r"^S\d{1,2}$", re.IGNORECASE)


def stable_id(*parts: str) -> str:
    h = hashlib.sha256()
    for p in parts:
        h.update(p.encode("utf-8", errors="ignore"))
        h.update(b"\0")
    return h.hexdigest()[:24]


def read_jsonl_ids(path: Path) -> Set[str]:
    if not path.exists():
        return set()
    done = set()
    with path.open("r", encoding="utf-8") as f:
        for line in f:
            line = line.strip()
            if not line:
                continue
            try:
                obj = json.loads(line)
                if "id" in obj:
                    done.add(obj["id"])
            except Exception:
                continue
    return done


def append_jsonl(path: Path, rows: List[Dict[str, Any]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("a", encoding="utf-8") as f:
        for r in rows:
            f.write(json.dumps(r, ensure_ascii=False) + "\n")


def load_tracks_mapping() -> Dict[str, str]:
    """
    Charge TRACKS depuis ENSTA_DATASET/config/tracks.py
    """
    import importlib.util

    tracks_py = CONFIG_DIR / "tracks.py"
    if not tracks_py.exists():
        raise FileNotFoundError(
            f"Mapping introuvable: {tracks_py}. Crée ENSTA_DATASET/config/tracks.py avec TRACKS = {{...}}"
        )

    spec = importlib.util.spec_from_file_location("tracks", str(tracks_py))
    module = importlib.util.module_from_spec(spec)
    assert spec and spec.loader
    spec.loader.exec_module(module)
    tracks = getattr(module, "TRACKS", None)
    if not isinstance(tracks, dict):
        raise ValueError("tracks.py doit contenir un dict TRACKS = {...}")
    # normalise clés
    return {str(k).strip().upper(): str(v).strip() for k, v in tracks.items()}


def build_merged_lookup(ws) -> Dict[str, str]:
    """
    Mappe chaque coordonnée (ex "B12") vers le coin haut-gauche de sa merged cell (ex "A12").
    """
    lookup: Dict[str, str] = {}
    for merged in ws.merged_cells.ranges:
        min_col, min_row, max_col, max_row = merged.bounds
        top_left = ws.cell(row=min_row, column=min_col).coordinate
        for r in range(min_row, max_row + 1):
            for c in range(min_col, max_col + 1):
                lookup[ws.cell(row=r, column=c).coordinate] = top_left
    return lookup


def row_text(ws, row_idx: int, merged_lookup: Dict[str, str]) -> str:
    """
    Récupère le texte "réel" d’une ligne, en respectant les merged cells.
    Déduplique les répétitions dues au remapping.
    """
    parts: List[str] = []
    # Bornes colonnes: openpyxl ws.max_column peut être grand si mise en forme.
    # On limite à 60 colonnes (A..BH), largement suffisant pour ces fiches.
    max_col = min(ws.max_column, 60)

    last = None
    for col in range(1, max_col + 1):
        cell = ws.cell(row=row_idx, column=col)
        coord = cell.coordinate
        top_left = merged_lookup.get(coord, coord)
        v = ws[top_left].value if top_left != coord else cell.value
        if v is None:
            continue
        s = str(v).strip()
        if not s:
            continue
        # évite répétitions
        if s != last:
            parts.append(s)
            last = s
    return " ".join(parts).strip()


def clean_desc(text: str) -> str:
    """
    Nettoyage minimal de la description.
    """
    t = (text or "").replace("\r\n", "\n").replace("\r", "\n")
    t = re.sub(r"[ \t]+", " ", t)
    t = re.sub(r"\n{3,}", "\n\n", t).strip()
    return t


def parse_path_metadata(fp: Path) -> Dict[str, Any]:
    """
    Déduit formation (FISE/FIPA), spécialité (code), semestre Sx depuis le chemin.
    """
    parts = [p for p in fp.parts]

    upper_parts = [p.upper() for p in parts]
    formation = None
    if any("FISE" in p for p in upper_parts):
        formation = "FISE"
    elif any("FIPA" in p for p in upper_parts):
        formation = "FIPA"

    semester = None
    sem_idx = None
    for i, p in enumerate(parts):
        if SEM_RE.match(p):
            semester = p.upper()
            sem_idx = i
            break

    # track_code = dossier juste avant semester (ex: ...\HYO\S4\file.xlsx)
    track_code = None
    if sem_idx is not None and sem_idx > 0:
        track_code = parts[sem_idx - 1].strip().upper()

    return {
        "formation": formation,
        "semester": semester,
        "track_code": track_code,
    }


def normalize_course_title(stem: str) -> str:
    """
    Nettoie le titre issu du nom de fichier.
    """
    s = stem.strip()
    s = re.sub(r"^fiche\s*cours\s*", "", s, flags=re.IGNORECASE)
    s = re.sub(r"\s+", " ", s).strip()
    return s


def extract_description_from_xlsx(fp: Path) -> Tuple[str, Optional[str]]:
    """
    Extrait description depuis rows 12 et 13, merged-aware.
    Retourne (desc, sheet_name_used).
    """
    wb = load_workbook(str(fp), data_only=True)
    try:
        ws = wb.active
        lookup = build_merged_lookup(ws)
        t12 = row_text(ws, ROW_DESC_1, lookup)
        t13 = row_text(ws, ROW_DESC_2, lookup)
        desc = "\n".join([x for x in [t12, t13] if x]).strip()
        return clean_desc(desc), ws.title
    finally:
        wb.close()


def main():
    tracks = load_tracks_mapping()

    OUT_JSONL.parent.mkdir(parents=True, exist_ok=True)
    OUT_JSONL.touch(exist_ok=True)

    done_ids = read_jsonl_ids(OUT_JSONL)

    # liste tous les fichiers xlsx de toutes les roots
    files: List[Tuple[Path, Path]] = []  # (root, file)
    for root in XLSX_ROOTS:
        if not root.exists():
            print(f"[WARN] root introuvable: {root}")
            continue
        for fp in root.rglob("*.xlsx"):
            if fp.name.startswith("~$"):  # fichier temp Excel
                continue
            files.append((root, fp))

    print(f"[INFO] XLSX trouvés: {len(files)} | déjà traités: {len(done_ids)}")
    if not files:
        print("❌ Aucun XLSX trouvé. Vérifie XLSX_ROOTS.")
        return

    kept = 0
    skipped_no_desc = 0
    unknown_tracks: Set[str] = set()
    missing_meta = 0

    buffer: List[Dict[str, Any]] = []

    for i, (root, fp) in enumerate(sorted(files, key=lambda x: str(x[1])), start=1):
        doc_id = stable_id("xlsx_course_sheet", str(fp.resolve()))
        if doc_id in done_ids:
            continue

        meta = parse_path_metadata(fp)
        if not meta.get("formation") or not meta.get("semester") or not meta.get("track_code"):
            missing_meta += 1

        track_code = (meta.get("track_code") or "").upper()
        track_label = tracks.get(track_code)
        if track_code and track_label is None:
            unknown_tracks.add(track_code)

        try:
            desc, sheet_name = extract_description_from_xlsx(fp)
        except Exception:
            continue

        if not desc or len(desc) < 40:
            skipped_no_desc += 1
            continue

        course_title = normalize_course_title(fp.stem)

        out_obj = {
            "id": doc_id,
            "source_type": "xlsx_course_sheet",
            "source": str(fp.resolve()),
            "title": course_title,
            "text": desc,
            "metadata": {
                "filename": fp.name,
                "relative_path": str(fp.relative_to(root)),
                "sheet_name": sheet_name,
                "formation": meta.get("formation"),
                "semester": meta.get("semester"),
                "track_code": track_code,
                "track_label": track_label,  # injecté via mapping manuel
            },
        }

        buffer.append(out_obj)
        done_ids.add(doc_id)
        kept += 1

        if len(buffer) >= 50:
            append_jsonl(OUT_JSONL, buffer)
            buffer = []

        if kept % PRINT_EVERY == 0:
            print(f"[OK] kept={kept} / scanned={i}")

    if buffer:
        append_jsonl(OUT_JSONL, buffer)

    print("\n✅ XLSX INGEST DONE")
    print(f"Output: {OUT_JSONL.resolve()}")
    print(f"Kept  : {kept} (nouveaux)")
    print(f"Skipped (no desc): {skipped_no_desc}")
    print(f"Missing path meta: {missing_meta}")

    if unknown_tracks:
        print("\n⚠️ Track codes inconnus (ajoute-les dans ENSTA_DATASET/config/tracks.py):")
        for t in sorted(unknown_tracks):
            print(" -", t)


if __name__ == "__main__":
    main()
